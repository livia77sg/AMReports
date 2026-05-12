"""Core processor: Track360 raw report → filled-out conversion templates."""

from __future__ import annotations

import io
import re
import shutil
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import config


@dataclass
class ProcessResult:
    """Summary of a processing run."""
    output_files: dict[str, bytes] = field(default_factory=dict)  # filename -> content
    report_date: str = ""
    total_signups: int = 0
    total_ftds: int = 0
    per_output_counts: dict[str, dict] = field(default_factory=dict)  # key -> {signups, ftd_lines}
    warnings: list[str] = field(default_factory=list)
    unmapped_brands: list[str] = field(default_factory=list)
    # Brands that ARE in brand_map.csv but missing a CPA in brand_values.csv.
    # Only applies to cpa-pattern outputs (HW/CW). Each entry is (raw_brand, vertical).
    missing_cpa_brands: list[tuple[str, str]] = field(default_factory=list)


def _normalize_brand_fallback(raw: str) -> str:
    """Fallback rule when brand isn't in brand_map.csv:
    lowercase, strip common geo/category suffixes, remove dashes/spaces.
    Example: 'Betfred-UK-Casino' -> 'betfred'.
    """
    if not raw:
        return ""
    s = raw.strip()
    # Strip trailing region/category fragments
    s = re.sub(r"-(UK|GB|US|EU|DE|IT|BR|AU|NL)-?(Casino|Bingo|Sport|Sports)?$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"-(Casino|Bingo|Sport|Sports)-(UK|GB|US|EU|DE|IT|BR|AU|NL)$", "", s, flags=re.IGNORECASE)
    s = s.lower().replace("-", "").replace(" ", "")
    return s


def _detect_category_fallback(raw: str) -> str:
    """Fallback category detection from raw brand name."""
    s = (raw or "").lower()
    if "bingo" in s:
        return "bingo"
    if "casino" in s:
        return "casino"
    return "sport"  # default


def load_brand_values(csv_path: Path) -> dict[tuple[str, str], int]:
    """Load brand_values.csv → {(clean_name_lower, vertical_lower): value}.

    Brand values are looked up by the clean (resolved) brand name and vertical
    (e.g. "hw" or "cw"). Per AM clarification, values are consistent across
    Google/Bing within a vertical, so no per-output key is needed here.
    """
    if not csv_path.exists():
        return {}
    df = pd.read_csv(csv_path).fillna("")
    out: dict[tuple[str, str], int] = {}
    for _, row in df.iterrows():
        clean = str(row["clean_name"]).strip().lower()
        vertical = str(row["vertical"]).strip().lower()
        try:
            value = int(round(float(row["value"])))
        except (ValueError, TypeError):
            continue
        out[(clean, vertical)] = value
    return out


def resolve_brand_value(
    clean_name: str,
    vertical: str,
    brand_values: dict[tuple[str, str], int],
) -> int | None:
    """Look up the fixed CPA value for a brand in a vertical. Returns None if
    not in the map — caller decides what to do (warn, default, skip)."""
    return brand_values.get((clean_name.strip().lower(), vertical.strip().lower()))


def load_brand_map(csv_path: Path) -> dict[tuple[str, str], tuple[str, str]]:
    """Load brand_map.csv → {(report_name_lower, output_key): (clean_name, category)}.

    `output_key` may be empty string ("") which means "default for all outputs".
    Lookup at resolve time first tries (name, specific_output_key), then falls
    back to (name, "")."""
    if not csv_path.exists():
        return {}
    df = pd.read_csv(csv_path).fillna("")
    out: dict[tuple[str, str], tuple[str, str]] = {}
    for _, row in df.iterrows():
        name = str(row["report_brand_name"]).strip().lower()
        ok = str(row.get("output_key", "")).strip().lower()
        clean = str(row["clean_name"]).strip()
        cat = str(row["category"]).strip().lower()
        out[(name, ok)] = (clean, cat)
    return out


def resolve_brand(
    raw: str,
    brand_map: dict[tuple[str, str], tuple[str, str]],
    output_key: str,
) -> tuple[str, str, bool]:
    """Return (clean_name, category, was_mapped).

    Resolution order:
      1. (raw_lower, output_key) — output-specific override
      2. (raw_lower, "") — global default
      3. heuristic fallback (lowercase + suffix strip)
    """
    if pd.isna(raw):
        return ("", "sport", False)
    key = str(raw).strip().lower()
    ok = output_key.strip().lower()
    if (key, ok) in brand_map:
        clean, cat = brand_map[(key, ok)]
        return (clean, cat, True)
    if (key, "") in brand_map:
        clean, cat = brand_map[(key, "")]
        return (clean, cat, True)
    return (_normalize_brand_fallback(str(raw)), _detect_category_fallback(str(raw)), False)


def _detect_report_date(df: pd.DataFrame) -> tuple[str, list[str]]:
    """Pick the most common date across Signup/FTD date columns.
    Returns (YYYY-MM-DD string, warnings)."""
    warnings = []
    dates = []
    for col in (config.COL_SIGNUP_DATE, config.COL_FTD_DATE):
        if col in df.columns:
            dates.extend(pd.to_datetime(df[col], errors="coerce").dropna().dt.date.tolist())
    if not dates:
        # Fallback: today (shouldn't happen with real reports)
        warnings.append("No dates found in report — used today's date as fallback.")
        return (datetime.now().strftime("%Y-%m-%d"), warnings)
    counter = Counter(dates)
    most_common, count = counter.most_common(1)[0]
    if len(counter) > 1:
        others = ", ".join(str(d) for d, _ in counter.most_common()[1:5])
        warnings.append(
            f"Report contains multiple dates. Used {most_common} ({count} rows). "
            f"Other dates: {others}."
        )
    return (most_common.strftime("%Y-%m-%d"), warnings)


def _build_rows_paired_ftd(
    df: pd.DataFrame,
    output_cfg: dict,
    report_date: str,
    brand_map: dict,
    unmapped_brands: set[str],
) -> list[list]:
    """Build rows for sport/bingo/casino-style outputs (paired_ftd pattern):
    - signups = 1 row at fixed 23:55, value=1
    - ftds = 2 rows at fixed 23:58 (brand + generic ftds), value=rounded revenue
    """
    site_ids = output_cfg["site_ids"]
    types = config.CHANNEL_TYPES[output_cfg["channel"]]

    sub = df[
        df[config.COL_SITE_ID].isin(site_ids)
        & df[config.COL_TYPE].astype(str).str.lower().isin([t.lower() for t in types])
    ].copy()

    if sub.empty:
        return []

    # Strip Bing prefix from Cid
    if output_cfg["channel"] == "bing":
        sub[config.COL_CID] = sub[config.COL_CID].astype(str).apply(
            lambda c: c[len(config.BING_CID_PREFIX):] if c.startswith(config.BING_CID_PREFIX) else c
        )

    collapse_casino = output_cfg["key"] in config.COLLAPSE_CASINO_OUTPUTS
    report_date_obj = pd.to_datetime(report_date).date()
    # Conversion Time as real datetime objects so Excel formats them properly
    sig_h, sig_m, sig_s = (int(x) for x in config.SIGNUP_TIME.split(":"))
    ftd_h, ftd_m, ftd_s = (int(x) for x in config.FTD_TIME.split(":"))
    signup_dt = datetime.combine(report_date_obj, datetime.min.time()).replace(
        hour=sig_h, minute=sig_m, second=sig_s
    )
    ftd_dt = datetime.combine(report_date_obj, datetime.min.time()).replace(
        hour=ftd_h, minute=ftd_m, second=ftd_s
    )
    rows: list[list] = []

    # === SIGNUP rows ===
    # Only include rows where the signup actually happened on the report date.
    # Stale signups from previous days that are still in the report should not
    # be re-counted today.
    signup_dates = pd.to_datetime(sub[config.COL_SIGNUP_DATE], errors="coerce").dt.date
    signup_df = sub[
        (sub[config.COL_SIGNUPS].fillna(0) > 0)
        & (signup_dates == report_date_obj)
    ]
    for _, r in signup_df.iterrows():
        cid = r[config.COL_CID]
        if pd.isna(cid) or str(cid).strip().lower() in ("na", "nan", ""):
            continue
        # If a row has signups > 1, emit one row per signup (matches manual workflow
        # where the AM clicks into the pivot and pulls one Cid per signup)
        n = int(r[config.COL_SIGNUPS])
        for _ in range(n):
            rows.append([str(cid), output_cfg["signup_label"], signup_dt, 1, config.CURRENCY])

    # === FTD rows ===
    # Same filter — only include FTDs that actually happened on the report date.
    # We collect brand-name rows and generic "ftds" rows separately so the final
    # output is: signups → all brand rows (or casino-collapse rows) → all ftds rows.
    ftd_dates = pd.to_datetime(sub[config.COL_FTD_DATE], errors="coerce").dt.date
    ftd_df = sub[
        (sub[config.COL_FTDS].fillna(0) > 0)
        & (ftd_dates == report_date_obj)
    ]
    brand_rows: list[list] = []
    ftds_rows: list[list] = []
    for _, r in ftd_df.iterrows():
        cid = r[config.COL_CID]
        if pd.isna(cid) or str(cid).strip().lower() in ("na", "nan", ""):
            continue
        n_ftds = int(r[config.COL_FTDS])
        revenue = r[config.COL_REVENUES]
        revenue_int = int(round(float(revenue))) if pd.notna(revenue) else 0
        # Per-FTD revenue split: if a row has multiple FTDs, divide and round
        per_ftd_rev = int(round(revenue_int / n_ftds)) if n_ftds > 0 else revenue_int

        clean, category, mapped = resolve_brand(r[config.COL_BRAND], brand_map, output_cfg["key"])
        if not mapped and pd.notna(r[config.COL_BRAND]):
            unmapped_brands.add(str(r[config.COL_BRAND]))

        for _ in range(n_ftds):
            if category == "casino" and collapse_casino:
                # Single row labeled as the output's casino_label.
                # Goes in the brand_rows bucket since it occupies the brand slot.
                brand_rows.append(
                    [str(cid), output_cfg["casino_label"], ftd_dt, per_ftd_rev, config.CURRENCY]
                )
            else:
                # Two-row treatment: brand row goes in brand_rows, generic ftds row at the end
                brand_rows.append(
                    [str(cid), f"{output_cfg['brand_prefix']}{clean}", ftd_dt, per_ftd_rev, config.CURRENCY]
                )
                ftds_rows.append(
                    [str(cid), output_cfg["ftd_label"], ftd_dt, per_ftd_rev, config.CURRENCY]
                )

    # Sort brand rows alphabetically by conversion name (column index 1).
    # Generic ftds rows stay grouped at the end in insertion order.
    brand_rows.sort(key=lambda r: str(r[1]).lower())
    rows.extend(brand_rows)
    rows.extend(ftds_rows)
    return rows


def _build_rows_cpa(
    df: pd.DataFrame,
    output_cfg: dict,
    report_date: str,
    brand_map: dict,
    brand_values: dict,
    unmapped_brands: set[str],
    missing_cpa: set[tuple[str, str]],
) -> list[list]:
    """Build rows for HW/CW-style outputs (cpa pattern):

    - Signups = 2 rows per signup, both at the row's actual Signup Date (real datetime):
        1. ("offline - {brand_clean}", value = brand CPA)
        2. ("offline - converted leads", value = brand CPA)
    - FTDs = 1 row per ftd at the row's actual FtdDate, labeled by sales_label,
      value = 1.

    Brand cleaning uses brand_map.csv (shared with paired_ftd outputs).
    Brand values come from brand_values.csv keyed on (clean_name, vertical).
    Brands not in either map are skipped with a warning.
    """
    site_ids = output_cfg["site_ids"]
    types = config.CHANNEL_TYPES[output_cfg["channel"]]
    vertical = output_cfg["vertical"]  # "hw" or "cw"

    sub = df[
        df[config.COL_SITE_ID].isin(site_ids)
        & df[config.COL_TYPE].astype(str).str.lower().isin([t.lower() for t in types])
    ].copy()
    if sub.empty:
        return []

    if output_cfg["channel"] == "bing":
        sub[config.COL_CID] = sub[config.COL_CID].astype(str).apply(
            lambda c: c[len(config.BING_CID_PREFIX):] if c.startswith(config.BING_CID_PREFIX) else c
        )

    signup_brand_rows: list[list] = []   # (offline - {brand}) rows
    signup_leads_rows: list[list] = []   # (offline - converted leads) rows
    sales_rows: list[list] = []          # (offline - sales) rows

    report_date_obj = pd.to_datetime(report_date).date()

    # --- Signups ---
    signup_dates_only = pd.to_datetime(sub[config.COL_SIGNUP_DATE], errors="coerce").dt.date
    signup_df = sub[
        (sub[config.COL_SIGNUPS].fillna(0) > 0)
        & (signup_dates_only == report_date_obj)
    ]
    for _, r in signup_df.iterrows():
        cid = r[config.COL_CID]
        if pd.isna(cid) or str(cid).strip().lower() in ("na", "nan", ""):
            continue
        raw_brand = r[config.COL_BRAND]
        if pd.isna(raw_brand):
            continue
        # Resolve brand via the shared brand_map
        clean_name, _category, mapped = resolve_brand(raw_brand, brand_map, output_cfg["key"])
        if not mapped:
            unmapped_brands.add(str(raw_brand))
        # Look up the CPA value
        cpa = resolve_brand_value(clean_name, vertical, brand_values)
        if cpa is None:
            # No CPA known — different problem from unmapped. Record it
            # separately so the UI can show a precise diagnostic.
            missing_cpa.add((str(raw_brand), vertical))
            continue
        # Use the actual signup datetime from the report, dropping seconds
        # to match how the AM displays times in her files (minute precision)
        sig_time = pd.to_datetime(r[config.COL_SIGNUP_DATE])
        if pd.isna(sig_time):
            continue
        sig_dt = sig_time.to_pydatetime().replace(second=0, microsecond=0)
        n = int(r[config.COL_SIGNUPS])
        for _ in range(n):
            signup_brand_rows.append(
                [str(cid), f"{output_cfg['brand_prefix']}{clean_name}", sig_dt, cpa, config.CURRENCY]
            )
            signup_leads_rows.append(
                [str(cid), output_cfg["signup_label"], sig_dt, cpa, config.CURRENCY]
            )

    # --- FTDs (sales) ---
    ftd_dates_only = pd.to_datetime(sub[config.COL_FTD_DATE], errors="coerce").dt.date
    ftd_df = sub[
        (sub[config.COL_FTDS].fillna(0) > 0)
        & (ftd_dates_only == report_date_obj)
    ]
    for _, r in ftd_df.iterrows():
        cid = r[config.COL_CID]
        if pd.isna(cid) or str(cid).strip().lower() in ("na", "nan", ""):
            continue
        ftd_time = pd.to_datetime(r[config.COL_FTD_DATE])
        if pd.isna(ftd_time):
            continue
        f_dt = ftd_time.to_pydatetime().replace(second=0, microsecond=0)
        n_ftds = int(r[config.COL_FTDS])
        for _ in range(n_ftds):
            sales_rows.append(
                [str(cid), output_cfg["sales_label"], f_dt, 1, config.CURRENCY]
            )

    # Order: brand-name signup rows first (alphabetical), then converted-leads
    # signup rows, then sales rows. AM's HW Google example shows: brand rows,
    # then converted-leads rows. Brand rows are sorted by conversion name.
    signup_brand_rows.sort(key=lambda r: str(r[1]).lower())
    rows: list[list] = []
    rows.extend(signup_brand_rows)
    rows.extend(signup_leads_rows)
    rows.extend(sales_rows)
    return rows


def _build_rows_for_output(
    df: pd.DataFrame,
    output_cfg: dict,
    report_date: str,
    brand_map: dict,
    brand_values: dict,
    unmapped_brands: set[str],
    missing_cpa: set[tuple[str, str]],
) -> list[list]:
    """Dispatcher that picks the right row-builder based on output pattern."""
    pattern = output_cfg.get("pattern", "paired_ftd")
    if pattern == "cpa":
        return _build_rows_cpa(
            df, output_cfg, report_date, brand_map, brand_values, unmapped_brands, missing_cpa
        )
    return _build_rows_paired_ftd(df, output_cfg, report_date, brand_map, unmapped_brands)


def _find_header_row(ws) -> int:
    """Locate the row that starts with 'Microsoft Click ID' or 'Google Click ID'.
    Data rows go directly below it. Returns 1-indexed row number."""
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        cell = ws.cell(row=row_idx, column=1).value
        if cell and isinstance(cell, str) and "Click ID" in cell:
            return row_idx
    raise ValueError("Could not find header row in template.")


def _write_template(
    template_path: Path,
    rows: list[list],
    output_path: Path,
    output_key: str,
):
    """Open template, clear existing example data rows, write new rows, save.

    Forces explicit number_format per column so output rows don't inherit weird
    formats from the template's example cells (e.g. scientific notation on Cid
    column, or stale date formats partway through). Also clears any background
    fill on data cells (Sport_Google template ships with blue fill on example
    rows; AM expects white).
    """
    from openpyxl.styles import PatternFill
    no_fill = PatternFill(fill_type=None)

    wb = load_workbook(template_path)
    ws = wb.active
    header_row = _find_header_row(ws)
    first_data_row = header_row + 1

    # Clear existing example data rows (templates ship with sample data).
    # We clear columns 1-6: 1-5 are the actual data columns, and column 6 is
    # cleared because some templates have annotation comments there (e.g.
    # "(same as brand's deal)") that the AMs don't include in real reports.
    last_existing = ws.max_row
    for r in range(first_data_row, last_existing + 1):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.number_format = "General"
            cell.fill = no_fill

    # Per-column number_format we explicitly set on every written row.
    # The [$-409] prefix on the time column forces US English locale so dates
    # always display as month/day/year, even if the viewer's Excel is set to
    # a European regional format (which would otherwise flip to dd/mm/yyyy).
    column_formats = {
        1: "General",                       # Cid
        2: "General",                       # Conversion Name
        3: "[$-409]m/d/yyyy h:mm;@",        # Conversion Time (locale-forced m/d/yyyy)
        4: "General",                       # Conversion Value
        5: "General",                       # Currency
    }

    for i, row_data in enumerate(rows):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=first_data_row + i, column=c)
            cell.value = val
            cell.number_format = column_formats[c]
            cell.fill = no_fill

    wb.save(output_path)


def process_report(
    report_bytes: bytes,
    templates_dir: Path,
    brand_map_path: Path,
    brand_values_path: Path | None = None,
) -> ProcessResult:
    """Main entry point. Read raw Track360 report bytes, return ProcessResult
    with all output files in memory."""
    result = ProcessResult()
    df = pd.read_excel(io.BytesIO(report_bytes))

    # Validate columns
    missing = [c for c in config.REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Required columns missing from report: {missing}")

    # Drop Cid == "na" (case-insensitive)
    cid_str = df[config.COL_CID].astype(str).str.strip().str.lower()
    pre = len(df)
    df = df[~cid_str.isin(["na", "nan", ""])].copy()
    skipped = pre - len(df)
    if skipped:
        result.warnings.append(f"Excluded {skipped} rows with Cid = 'na'.")

    # Detect report date
    report_date, date_warnings = _detect_report_date(df)
    result.report_date = report_date
    result.warnings.extend(date_warnings)

    # Load brand map + per-vertical brand values (CPA table)
    brand_map = load_brand_map(brand_map_path)
    brand_values = load_brand_values(brand_values_path) if brand_values_path else {}
    unmapped: set[str] = set()
    missing_cpa: set[tuple[str, str]] = set()

    # Process each output
    import tempfile
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        for output_cfg in config.OUTPUTS:
            template_path = templates_dir / output_cfg["template"]
            if not template_path.exists():
                result.warnings.append(f"Missing template: {template_path.name}")
                continue

            rows = _build_rows_for_output(
                df, output_cfg, report_date, brand_map, brand_values, unmapped, missing_cpa
            )

            # Count for summary — different patterns count differently.
            # paired_ftd: signup_label rows are signups; everything else is ftds.
            # cpa: signup_label + brand rows (when at signup time) are signups;
            #      sales_label rows are ftds.
            pattern = output_cfg.get("pattern", "paired_ftd")
            if pattern == "cpa":
                sales_label = output_cfg.get("sales_label", "")
                ftd_count = sum(1 for r in rows if r[1] == sales_label)
                sig_count = len(rows) - ftd_count
            else:
                sig_label = output_cfg.get("signup_label", "")
                sig_count = sum(1 for r in rows if r[1] == sig_label)
                ftd_count = len(rows) - sig_count
            result.per_output_counts[output_cfg["key"]] = {
                "signups": sig_count,
                "ftd_rows": ftd_count,
            }
            result.total_signups += sig_count
            result.total_ftds += ftd_count

            # Skip empty outputs — no point in producing an xlsx with zero data rows
            if not rows:
                continue

            out_filename = f"{output_cfg['output_prefix']}_{report_date}.xlsx"
            out_path = tmp / out_filename
            _write_template(template_path, rows, out_path, output_cfg["key"])
            result.output_files[out_filename] = out_path.read_bytes()

    result.unmapped_brands = sorted(unmapped)
    if unmapped:
        result.warnings.append(
            f"{len(unmapped)} brand(s) not in brand_map.csv (default rule was used): "
            f"{', '.join(sorted(unmapped))}. "
            f"Add them to brand_map.csv for full control."
        )

    result.missing_cpa_brands = sorted(missing_cpa)
    if missing_cpa:
        # Group by vertical for a cleaner message
        by_vertical: dict[str, list[str]] = {}
        for raw, vert in missing_cpa:
            by_vertical.setdefault(vert, []).append(raw)
        lines = []
        for vert in sorted(by_vertical):
            lines.append(f"  {vert.upper()}: {', '.join(sorted(by_vertical[vert]))}")
        result.warnings.append(
            f"{len(missing_cpa)} brand(s) missing a CPA value in brand_values.csv — "
            f"these rows were SKIPPED. Add them to brand_values.csv with the correct value:\n"
            + "\n".join(lines)
        )

    return result